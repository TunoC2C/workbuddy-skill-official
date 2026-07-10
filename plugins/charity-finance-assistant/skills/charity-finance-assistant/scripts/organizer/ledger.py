"""
v1.3 台账模块：生成 Excel（多 Sheet）+ CSV 兜底。

拆解策略：
- _make_styles()           样式工厂
- _write_header()          写标题 + 表头
- _write_data_rows()       写数据
- _write_total_row()       写合计行
- _set_column_widths()     列宽设置
- generate_ledger_xlsx()   只编排（圈复杂度 ≈ 4）
"""
import csv
from datetime import datetime
from pathlib import Path

from organizer.config import LEDGER_COLS

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


def _make_styles():
    """工厂方法：返回所有需要的样式对象"""
    return {
        "header_font": Font(bold=True, color="FFFFFF", size=11),
        "header_fill": PatternFill("solid", fgColor="1F4E78"),
        "border": Border(
            left=Side(style="thin", color="BFBFBF"),
            right=Side(style="thin", color="BFBFBF"),
            top=Side(style="thin", color="BFBFBF"),
            bottom=Side(style="thin", color="BFBFBF"),
        ),
        "center": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "left_wrap": Alignment(horizontal="left", vertical="center", wrap_text=True),
        "title_font": Font(bold=True, size=13),
        "highlight_fill": PatternFill("solid", fgColor="FFF2CC"),
        "total_fill": PatternFill("solid", fgColor="D9E1F2"),
        "bold_font": Font(bold=True),
    }


def _write_header(ws, styles):
    """写标题行 + 表头行"""
    ws.cell(
        row=1, column=1,
        value=f"发票台账（自动生成 {datetime.now():%Y-%m-%d %H:%M}）",
    )
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(LEDGER_COLS))
    ws.cell(row=1, column=1).font = styles["title_font"]
    ws.cell(row=1, column=1).alignment = styles["center"]

    for i, col in enumerate(LEDGER_COLS, 1):
        c = ws.cell(row=2, column=i, value=col)
        c.font = styles["header_font"]
        c.fill = styles["header_fill"]
        c.alignment = styles["center"]
        c.border = styles["border"]


_LEFT_WRAP_COLS = {"购买方/捐赠方", "销售方/受赠方", "货物或服务项目", "备注"}


def _coerce_cell_value(col, value):
    """价税合计列尝试转 float"""
    if col != "价税合计(元)" or not value:
        return value
    try:
        return float(value)
    except (ValueError, TypeError):
        return value


def _write_data_rows(ws, organized, styles):
    """写所有数据行（圈复杂度 ≈ 4）"""
    for row_idx, item in enumerate(organized, 3):
        for col_idx, col in enumerate(LEDGER_COLS, 1):
            value = _coerce_cell_value(col, item.get(col, ""))
            c = ws.cell(row=row_idx, column=col_idx, value=value)
            c.alignment = styles["left_wrap"] if col in _LEFT_WRAP_COLS else styles["center"]
            c.border = styles["border"]
            if item.get("数据状态") == "待人工补录":
                c.fill = styles["highlight_fill"]


def _is_numeric(s):
    """判断字符串可转为数值（用于合计）"""
    return s and str(s).replace(".", "").replace("-", "").isdigit()


def _write_total_row(ws, organized, styles):
    """写合计行"""
    total_row = len(organized) + 3
    ws.cell(row=total_row, column=1, value="合计").font = styles["bold_font"]
    total_sum = sum(
        float(x["价税合计(元)"])
        for x in organized
        if _is_numeric(x.get("价税合计(元)"))
    )
    cell = ws.cell(row=total_row, column=12, value=total_sum)
    cell.font = styles["bold_font"]
    for i in range(1, len(LEDGER_COLS) + 1):
        ws.cell(row=total_row, column=i).fill = styles["total_fill"]
        ws.cell(row=total_row, column=i).border = styles["border"]


def _set_column_widths(ws):
    """统一设置列宽和首行高度"""
    widths = [6, 18, 22, 16, 14, 12, 28, 28, 32, 12, 12, 14, 24, 22, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[2].height = 30


def generate_ledger_xlsx(organized, output_folder):
    """
    生成 Excel 台账（15 列 + 合计行）。
    圈复杂度 = 2：HAS_OPENPYXL 检查 + 编排。
    """
    if not HAS_OPENPYXL:
        return None

    output = Path(output_folder)
    ledger_path = output / "票据台账_自动生成.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "发票台账"

    styles = _make_styles()
    _write_header(ws, styles)
    _write_data_rows(ws, organized, styles)
    _write_total_row(ws, organized, styles)
    _set_column_widths(ws)
    ws.freeze_panes = "A3"

    wb.save(str(ledger_path))
    return ledger_path


def generate_ledger_csv(organized, output_folder):
    """CSV 兜底台账（圈复杂度 = 2）"""
    output = Path(output_folder)
    ledger_path = output / "票据台账_自动生成.csv"
    with open(str(ledger_path), "w", newline="", encoding="utf-8-sig") as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=LEDGER_COLS)
        writer.writeheader()
        for item in organized:
            writer.writerow({k: item.get(k, "") for k in LEDGER_COLS})
    return ledger_path
